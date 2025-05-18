import utils_1 as u1
import surveyor_method as sm
import shapely as sp  
from pathlib import Path
import openpyxl
from openpyxl.styles import Border, Side
from typing import List, Tuple, Union

def process_autocad_log(log_file_path) -> Tuple[str, List[sp.Polygon], sp.GeometryCollection]:
    """
    AutoCADのログファイルから座標を抽出し、ポリゴンリストとGeometryCollectionを作成する。
    """

    path_obj_file: str = Path(log_file_path) 
    base_file_name: str = path_obj_file.stem

    parsers: List[List[Tuple[float, float]]] = u1.parse_autocad_log_for_coordinates(log_file_path)
    u1.append_first_to_end_if_mismatched(parsers)

    polygon_list: List[sp.Polygon] = [u1.create_polygon_from_xy_coords(l) for l in parsers]
    collection = sp.GeometryCollection(polygon_list)

    return base_file_name, polygon_list, collection

def calculate_mesh_overlaps(polygon_list, collection, bounds_offset=10):
    """
    ポリゴンとメッシュの重複を計算し、関連データを整理する。
    """
    collection :sp.GeometryCollection = collection
    caluculatin_area = u1.get_calculation_area(collection.bounds, bounds_offset)
    grid_aria = u1.GridArea(caluculatin_area)
    check_MeshList = u1.check_MeshList_polygon_overlap(grid_aria.mesh_list, polygon_list)

    check_MeshList_of_not_100_empty = [
        item for item in check_MeshList if item[0] == "partial_overlap" or item[0] == "multipolygon"
    ]
    check_MeshList_of_all_ovelap = [
        [item[2].index, 100] for item in check_MeshList if item[0] == "all_overlap"
    ]

    csv_of_surveyor_method = []
    
    for check in check_MeshList_of_not_100_empty:
        check: tuple[str, Union[sp.Polygon, sp.MultiPolygon, None], u1.MeshCell] 
        if check[0] != "None":
            if check[0] != "multipolygon":
                csv_of_surveyor_method.append([check[2].index,
                                sm.calculate_polygon_area_surveyor_method(check[1])])
            elif check[0] == "multipolygon":
                for i in check[1].geoms:
                    csv_of_surveyor_method.append([check[2].index,
                                sm.calculate_polygon_area_surveyor_method(i)])

    csv_of_main_list = [[sub_list[0], sub_list[1]["area"]] for sub_list in csv_of_surveyor_method]
    for i in check_MeshList_of_all_ovelap:
        csv_of_main_list.append(i)

    sorted_csv_main_list = sorted(csv_of_main_list, key=lambda item: (item[0][0], item[0][1]))

    area_from_main_list = sum(item[1] for item in sorted_csv_main_list)
    print(area_from_main_list, ":10mメッシュの積上げ")
    print(collection.area, ":入力したポリゴンの積上げ面積")
    print(f"{area_from_main_list - collection.area:.10f}:差")

    return csv_of_surveyor_method, sorted_csv_main_list

def create_excel_report(base_file_name, csv_of_surveyor_method, sorted_csv_main_list, template_excel_filename, output_excel_filename_template):
    """
    計算結果をExcelファイルに出力する。
    """
    output_excel_filename = output_excel_filename_template.format(base_file_name=base_file_name)
    wb = openpyxl.load_workbook(template_excel_filename)

    # --- エクセル書式設定 --------
    thin_border_side = Side(style='thin', color='000000')
    every_side_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side
    )
    # --------------------------

    # 「倍面積法」シートの作成
    sheet_name_surveyor = '倍面積法'
    if sheet_name_surveyor in wb.sheetnames:
        ws_surveyor = wb[sheet_name_surveyor]
    else:
        ws_surveyor = wb.create_sheet(title=sheet_name_surveyor)

    current_row_surveyor = 1
    ws_surveyor.cell(row=current_row_surveyor, column=1, value=base_file_name)
    current_row_surveyor += 1
    start_col_surveyor = 1

    for item in csv_of_surveyor_method:
        ws_surveyor.cell(row=current_row_surveyor, column=start_col_surveyor, value="index_x")
        ws_surveyor.cell(row=current_row_surveyor, column=start_col_surveyor + 1, value="index_y")
        ws_surveyor.cell(row=current_row_surveyor, column=start_col_surveyor + 2, value=item[0][0])
        ws_surveyor.cell(row=current_row_surveyor, column=start_col_surveyor + 3, value=item[0][1])
        current_row_surveyor += 1

        header_row_1 = ["Xn", "Yn", "Yn+1 - Yn-1", "Xn * (Yn+1 - Yn-1)"]
        for col_idx, header_value in enumerate(header_row_1, start=1):
            ws_surveyor.cell(row=current_row_surveyor, column=col_idx, value=header_value)
        current_row_surveyor += 1

        list_of_coords = item[1]["data_list"]
        for coord_list in list_of_coords:
            for col_idx, cell_value in enumerate(coord_list, start=start_col_surveyor):
                ws_surveyor.cell(row=current_row_surveyor, column=col_idx, value=cell_value)
            current_row_surveyor += 1

        value1 = item[1]["double_area"]
        ws_surveyor.cell(row=current_row_surveyor, column=3, value="倍面識")
        ws_surveyor.cell(row=current_row_surveyor, column=4, value=value1)
        current_row_surveyor += 1

        value2 = item[1]["area"]
        ws_surveyor.cell(row=current_row_surveyor, column=3, value="面識")
        ws_surveyor.cell(row=current_row_surveyor, column=4, value=value2)
        current_row_surveyor += 2 # Add an empty line

    for row_idx in range(1, current_row_surveyor -1): # Adjusted range for border
        for col_idx in range(1, 5):
            cell = ws_surveyor.cell(row=row_idx, column=col_idx)
            cell.border = every_side_border
    ws_surveyor.oddFooter.center.text = "&P / &N"

    # 「一覧表」シートの作成
    sheet_name_list = '一覧表'
    if sheet_name_list in wb.sheetnames:
        ws_list = wb[sheet_name_list]
    else:
        ws_list = wb.create_sheet(title=sheet_name_list)

    current_row_list = 1
    ws_list.cell(row=current_row_list, column=2, value=base_file_name)
    current_row_list += 3
    ws_list.cell(row=current_row_list, column=4, value=f"=SUM(D5:D{len(sorted_csv_main_list) + current_row_list})") # Corrected SUM range
    current_row_list += 1

    for item in sorted_csv_main_list:
        ws_list.cell(row=current_row_list, column=1, value=f'="x"&B{current_row_list}&"y"&C{current_row_list}')
        ws_list.cell(row=current_row_list, column=2, value=item[0][0])
        ws_list.cell(row=current_row_list, column=3, value=item[0][1])
        ws_list.cell(row=current_row_list, column=4, value=item[1])
        ws_list.cell(row=current_row_list, column=5, value=f'=IFERROR(VLOOKUP(A{current_row_list},健全度!$A$1:$B$18000,2,FALSE),1)')
        current_row_list += 1

    ws_list.print_area = f'B1:H{current_row_list - 1}'
    ws_list.oddFooter.center.text = "&P / &N"

    for row_idx in range(5, len(sorted_csv_main_list) + 5):
        for col_idx in range(2, 9):
            cell = ws_list.cell(row=row_idx, column=col_idx)
            cell.border = every_side_border

    wb.save(output_excel_filename)
    print(f"レポートが {output_excel_filename} に保存されました。")

def main():
    """
    メイン処理を実行する関数。
    """
    log_file_path = "./test/10mメッシュに複数のポリゴンがあるテスト用.log"
    template_excel_filename = "./output/様式_操業面積.xlsx"
    output_excel_filename_template = "./output/{base_file_name}.xlsx" # テンプレート文字列を使用

    base_file_name, polygon_list, collection = process_autocad_log(log_file_path)
    csv_of_surveyor_method, sorted_csv_main_list = calculate_mesh_overlaps(polygon_list, collection)
    create_excel_report(base_file_name, csv_of_surveyor_method, sorted_csv_main_list, template_excel_filename, output_excel_filename_template)

if __name__ == "__main__":
    main()