from utils_1 import *
from surveyor_method import *
from shapely import *
from pathlib import Path
import csv
import openpyxl
from openpyxl.styles import Border, Side

log_file_path = "./test/autocad_log_9.log" 
path_obj_file = Path(log_file_path)
base_file_name = path_obj_file.stem

# autocad のログファイルから x,yの座標を抽出する
# args: 配列の配列
parsers = u1.parse_autocad_log_for_coordinates(log_file_path)

# x,yの座標配列の最初と最後が違うならば、最初を最後に追加する
append_first_to_end_if_mismatched(parsers)

# ポリゴンの配列をを作る
# 例：<POLYGON ((48.237 36.228, 47.407 43.922, 56.411 43.803, ...>, 
#   　<POLYGON ((22.94 12.662, 7.33 14.378, 5.406 24.502, ...>]
polygon_list = []
for l in parsers:
    polygon_list.append(create_polygon_from_xy_coords(l))

# ポリゴンの組をGEOMETRYCOLLECTIONにする
collection = GeometryCollection(polygon_list)

# 計算範囲の領域(bounds)を整数のタプルで出す
caluculatin_area = get_calculation_area(collection.bounds, 10)

# 計算エリアをGridAreaクラスに渡す
grid_aria = GridArea(caluculatin_area)

# (flag, 重複したポリゴン, 10*10メッシュのインスタンス)
check_MeshList = check_MeshList_polygon_overlap(grid_aria.mesh_list, polygon_list)

# Noneとall_overlap以外
check_MeshList_of_not_100_empty = [item for item in check_MeshList if item[0] == "partial_overlap" or item[0] == "multipolygon"]

# all_overlapだけ
check_MeshList_of_all_ovelap = [[item[2].index, 100] for item in check_MeshList if item[0] == "all_overlap"]

# csv_of_surveyor_method
# [ 
#   [(index_x,index_y), [倍面積過程], 倍面積, 面積],[ ] ... 
# ]

csv_of_surveyor_method = []
for check in check_MeshList_of_not_100_empty:
    if check[0] != "None":
        if check[0] != "multipolygon":
            csv_of_surveyor_method.append([check[2].index,
                            calculate_polygon_area_surveyor_method(check[1])])
        elif check[0] == "multipolygon":
            # print(check[1].geoms)
            for i in check[1].geoms:
                # print(i)
                csv_of_surveyor_method.append([check[2].index,
                            calculate_polygon_area_surveyor_method(i)])

# print(csv_of_surveyor_method[0])

# メインリストの作成にあたり、100とNone以外のリストを作る
csv_of_main_list = [[sub_list[0], sub_list[1]["area"]] for sub_list in csv_of_surveyor_method]

# メインリストに100を合成
for i in check_MeshList_of_all_ovelap:
    csv_of_main_list.append(i)



# indexでソートしたリスト
sorted_csv_main_list = sorted(csv_of_main_list, key = lambda item: (item[0][0], item[0][1]))

area_from_main_list = sum(item[1] for item in sorted_csv_main_list)
print(area_from_main_list, ":10mメッシュの積上げ")
print(collection.area, ":入力したポリゴンの積上げ面積")
print(f"{area_from_main_list-collection.area:.10f}:差")

# --- エクセル書式設定 --------

thin_border_side = Side(style='thin', color='000000')
every_side_border = Border(
    left=thin_border_side,
    right=thin_border_side,
    top=thin_border_side,
    bottom=thin_border_side
)

# --------------------------

template_excel_filename = "./output/様式_操業面積.xlsx"
output_excel_filename = f"./output/{base_file_name}.xlsx" 

wb = openpyxl.load_workbook(template_excel_filename)

sheet_name = '倍面積法'
if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
else:
    ws = wb.create_sheet(title=sheet_name) 

current_row = 1 
ws.cell(row=current_row, column=1, value=base_file_name)

current_row += 1
start_col = 1

for item in csv_of_surveyor_method:
    # 1. item[0] のタプルを書き出す
    ws.cell(row=current_row, column=start_col, value="index_x")
    ws.cell(row=current_row, column=start_col + 1, value="index_y")
    ws.cell(row=current_row, column=start_col + 2, value=item[0][0])
    ws.cell(row=current_row, column=start_col + 3, value=item[0][1])
    current_row += 1

    # ヘッダー行
    header_row_1 = ["Xn", "Yn", "Yn+1 - Yn-1", "Xn * (Yn+1 - Yn-1)"]
    for col_idx, header_value in enumerate(header_row_1, start=1):
        ws.cell(row=current_row, column=col_idx, value=header_value)
    current_row += 1

    # 2. データをループして書く
    list_of_coords = item[1]["data_list"]
    for coord_list in list_of_coords:
        for col_idx, cell_value in enumerate(coord_list, start=start_col):
            ws.cell(row=current_row, column=col_idx, value=cell_value)
        current_row += 1

    # 3. 「倍面識, 値」として書き出す
    value1 = item[1]["double_area"]
    ws.cell(row=current_row, column=3, value="倍面識")
    ws.cell(row=current_row, column=4, value=value1)
    ws.cell(row=current_row, column=1, value=None)
    ws.cell(row=current_row, column=2, value=None)
    current_row += 1

    # 4. 「面識, 値」として書き出す
    value2 = item[1]["area"]
    ws.cell(row=current_row, column=3, value="面識")
    ws.cell(row=current_row, column=4, value=value2)
    ws.cell(row=current_row, column=1, value=None)
    ws.cell(row=current_row, column=2, value=None)
    current_row += 1

    # 各itemの区切りとして空行を入れる場合 (任意)
    current_row += 1

# --- 罫線入力とフッター　----------------------

for row_idx in range(1, current_row - 1): 
    for col_idx in range(1, 5): 
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = every_side_border

ws.oddFooter.center.text = "&P / &N"



# --------一覧表-------------------------------
sheet_name = '一覧表' 
if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
else:
    ws = wb.create_sheet(title=sheet_name) 

current_row = 1

ws.cell(row=current_row, column=2, value=base_file_name)

current_row += 3

ws.cell(row=current_row, column=4, value=f"=sum(D5:D{len(sorted_csv_main_list) + current_row})")
current_row += 1

for item in sorted_csv_main_list:
    ws.cell(row=current_row, column=1, value=f'="x"&B{current_row}&"y"&C{current_row}')
    ws.cell(row=current_row, column=2, value=item[0][0])
    ws.cell(row=current_row, column=3, value=item[0][1])
    ws.cell(row=current_row, column=4, value=item[1])
    ws.cell(row=current_row, column=5, value=f'=IFERROR(VLOOKUP(A{current_row},健全度!$A$1:$B$18000,2,FALSE),1)')

    current_row += 1

# 変更を新しいファイルとして保存 (元のテンプレートファイルは変更されない)

# ws.page_setup.fitToWidth = 0
ws.print_area = f'B1:H{current_row - 1}'

ws.oddFooter.center.text = "&P / &N"

# --- 罫線入力　----------------------

for row_idx in range(5, len(sorted_csv_main_list)+5): 
    for col_idx in range(2, 9): 
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = every_side_border

wb.save(output_excel_filename)










